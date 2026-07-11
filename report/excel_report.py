import os
import json
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

FILE_NAME = "reports/TURION_AI_Trader.xlsx"
PORTFOLIO_FILE = "reports/paper_portfolio.json"


def save_market_summary(
    symbol,
    timeframe,
    price,
    ema20,
    ema50,
    rsi,
    market_state,
    market_structure,
    action,
    support=None,
    resistance=None,
    atr=None,
    volume_ratio=None,
    candle_pattern=None,
    ai_decision=None,
    ai_confidence=None
):

    os.makedirs("reports", exist_ok=True)

    # Create Excel File
    if not os.path.exists(FILE_NAME):

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Market Summary"

        sheet.append([
            "Run ID",
            "Date",
            "Time",
            "Symbol",
            "Timeframe",
            "Price",
            "EMA20",
            "EMA50",
            "RSI",
            "Market State",
            "Market Structure",
            "AI Decision",
            "Support",
            "Resistance",
            "ATR14",
            "Volume Ratio",
            "Candle Pattern",
            "AI Engine Decision",
            "AI Confidence %"
        ])

        # ---------------- Header Style ----------------

        blue_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        white_font = Font(
            bold=True,
            color="FFFFFF"
        )

        center = Alignment(
            horizontal="center",
            vertical="center"
        )

        thin = Side(
            border_style="thin",
            color="000000"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        for cell in sheet[1]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border

        # Column Width

        widths = {
            "A":10,
            "B":14,
            "C":12,
            "D":12,
            "E":12,
            "F":12,
            "G":12,
            "H":12,
            "I":10,
            "J":18,
            "K":18,
            "L":18,
            "M":12,
            "N":12,
            "O":10,
            "P":12,
            "Q":18,
            "R":16,
            "S":14
        }

        for col, width in widths.items():
            sheet.column_dimensions[col].width = width

        # Freeze Header
        sheet.freeze_panes = "A2"

        # Filter
        sheet.auto_filter.ref = "A1:S1"

        workbook.save(FILE_NAME)

    workbook = load_workbook(FILE_NAME)

    sheet = workbook["Market Summary"]

    # Updated: 2026-07-11 - backfill Support/Resistance/ATR14 headers for report files created before these columns existed
    if sheet["M1"].value != "Support":
        sheet["M1"] = "Support"

    if sheet["N1"].value != "Resistance":
        sheet["N1"] = "Resistance"

    if sheet["O1"].value != "ATR14":
        sheet["O1"] = "ATR14"

    if sheet["P1"].value != "Volume Ratio":
        sheet["P1"] = "Volume Ratio"

    if sheet["Q1"].value != "Candle Pattern":
        sheet["Q1"] = "Candle Pattern"

    if sheet["R1"].value != "AI Engine Decision":
        sheet["R1"] = "AI Engine Decision"

    if sheet["S1"].value != "AI Confidence %":
        sheet["S1"] = "AI Confidence %"

    run_id = sheet.max_row

    now = datetime.now()

    sheet.append([

        run_id,

        now.strftime("%d-%m-%Y"),

        now.strftime("%H:%M:%S"),

        symbol,

        timeframe,

        round(price, 2),

        round(ema20, 2),

        round(ema50, 2),

        round(rsi, 2),

        market_state,

        market_structure,

        action,

        round(support, 2) if support is not None else "",

        round(resistance, 2) if resistance is not None else "",

        round(atr, 2) if atr is not None else "",

        round(volume_ratio, 2) if volume_ratio is not None else "",

        candle_pattern if candle_pattern is not None else "",

        ai_decision if ai_decision is not None else "",

        round(ai_confidence, 1) if ai_confidence is not None else ""

    ])

    # ---------------- Color Coding ----------------

    green = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    yellow = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
        end_color="FFF2CC"
    )

    gray = PatternFill(
        fill_type="solid",
        start_color="D9D9D9",
        end_color="D9D9D9"
    )

    last_row = sheet.max_row

    state_cell = sheet[f"J{last_row}"]
    structure_cell = sheet[f"K{last_row}"]
    decision_cell = sheet[f"L{last_row}"]

    # Market State

    if "Bullish" in str(state_cell.value):
        state_cell.fill = green

    elif "Bearish" in str(state_cell.value):
        state_cell.fill = red

    else:
        state_cell.fill = yellow

    # Market Structure

    if "Bullish" in str(structure_cell.value):
        structure_cell.fill = green

    elif "Bearish" in str(structure_cell.value):
        structure_cell.fill = red

    else:
        structure_cell.fill = yellow

    # AI Decision

    value = str(decision_cell.value).upper()

    if "BUY" in value:
        decision_cell.fill = green

    elif "SELL" in value:
        decision_cell.fill = red

    elif "WAIT" in value:
        decision_cell.fill = yellow

    else:
        decision_cell.fill = gray

    update_dashboard(workbook, sheet)

    workbook.save(FILE_NAME)

    print("Excel Database Updated Successfully.")


def update_dashboard(workbook, market_sheet):

    if "Dashboard" in workbook.sheetnames:

        dashboard = workbook["Dashboard"]
        dashboard.delete_rows(1, dashboard.max_row)
        dashboard._charts = []

    else:

        dashboard = workbook.create_sheet("Dashboard", 0)

    blue_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    label_font = Font(bold=True)
    section_font = Font(bold=True, size=12)

    dashboard["A1"] = "TURION AI TRADER - DASHBOARD"
    dashboard["A1"].font = title_font
    dashboard["A1"].fill = blue_fill
    dashboard.merge_cells("A1:D1")

    last_row = market_sheet.max_row

    def latest(col):
        return market_sheet[f"{col}{last_row}"].value

    kpis = [
        ("Last Updated", f"{latest('B')} {latest('C')}"),
        ("Symbol", latest("D")),
        ("Current Price", latest("F")),
        ("Market State", latest("J")),
        ("Market Structure", latest("K")),
        ("AI Decision", latest("L")),
        ("Support", latest("M")),
        ("Resistance", latest("N")),
        ("ATR14", latest("O")),
        ("Candle Pattern", latest("Q")),
        ("AI Engine Decision", latest("R")),
        ("AI Confidence %", latest("S"))
    ]

    row = 3

    for label, value in kpis:

        dashboard[f"A{row}"] = label
        dashboard[f"A{row}"].font = label_font
        dashboard[f"B{row}"] = value

        row += 1

    row += 1

    dashboard[f"A{row}"] = "PAPER TRADING"
    dashboard[f"A{row}"].font = section_font

    row += 1

    if os.path.exists(PORTFOLIO_FILE):

        with open(PORTFOLIO_FILE, "r") as f:
            portfolio = json.load(f)

        closed = portfolio["Closed Trades"]
        total_pnl = sum(t["PnL"] for t in closed)

        paper_kpis = [
            ("Cash", round(portfolio["Cash"], 2)),
            ("Open Position", "Yes" if portfolio["Position"] else "No"),
            ("Closed Trades", len(closed)),
            ("Total PnL", round(total_pnl, 2))
        ]

        for label, value in paper_kpis:

            dashboard[f"A{row}"] = label
            dashboard[f"A{row}"].font = label_font
            dashboard[f"B{row}"] = value

            row += 1

    else:

        dashboard[f"A{row}"] = "No paper trading data yet"

        row += 1

    dashboard.column_dimensions["A"].width = 20
    dashboard.column_dimensions["B"].width = 22

    # Price Trend Chart over the recent runs
    max_points = 30
    start = max(2, last_row - max_points + 1)

    if last_row >= 2:

        chart = LineChart()
        chart.title = "Price Trend (Recent Runs)"
        chart.y_axis.title = "Price"
        chart.x_axis.title = "Run"
        chart.width = 18
        chart.height = 9

        data_ref = Reference(market_sheet, min_col=6, min_row=start, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=False)

        dashboard.add_chart(chart, "D3")